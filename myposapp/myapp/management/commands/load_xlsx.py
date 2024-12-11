import os
from django.apps import apps
from django.db.models import *
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from myapp.models import *


class Command(BaseCommand):
    help = '''Load data from an Excel file into the database.
    Usage: python manage.py load_xlsx -i myapp/fixtures/DATA.xlsx
    '''

    def add_arguments(self, parser):
        parser.add_argument('-i', '--input', type=str, required=True, help='Path to the Excel file')
        parser.add_argument('-s', '--skip_id', action='store_true', help='Skip ID fields in data loading')

    def handle(self, *args, **options):
        # Validate and load the Excel file
        xlsx_path = options['input']
        if not os.path.exists(xlsx_path):
            self.stderr.write(f"Error: File '{xlsx_path}' not found.")
            return

        try:
            workbook = load_workbook(xlsx_path)
            self.stdout.write(f"Loaded workbook: {xlsx_path}")
        except Exception as e:
            self.stderr.write(f"Error loading Excel file: {e}")
            return

        # Specify the models to load data for
        target_models = ['Category', 'Product', 'Option']
        app_config = apps.get_app_config('myapp')

        for model_name in target_models:
            if model_name not in workbook.sheetnames:
                self.stdout.write(f"Sheet '{model_name}' not found in the Excel file. Skipping.")
                continue

            model = app_config.get_model(model_name)
            sheet = workbook[model_name]
            headers = None

            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_index == 1:
                    # Set headers from the first row
                    headers = [col.lower() if col else None for col in row]
                    headers = [col for col in headers if col]  # Remove None values
                    self.stdout.write(f"Headers for '{model_name}': {headers}")
                    continue

                # Map row data to model fields
                data = {headers[i]: row[i] for i in range(len(headers)) if headers[i] in [f.name for f in model._meta.fields]}
                if options['skip_id']:
                    data.pop('id', None)

                # Clean and prepare data for insertion
                data = {k: (v.strip() if isinstance(v, str) else v) for k, v in data.items() if v is not None}

                # Handle special field types
                for field_name, value in data.items():
                    field = model._meta.get_field(field_name)
                    try:
                        if isinstance(field, ForeignKey) and value:
                            related_model = field.related_model
                            if field_name == 'category':  # For Product model, category is a ForeignKey
                                related_obj = related_model.objects.get(name=value)  # Assuming 'name' is unique
                            elif field_name == 'product':  # For Option model, product is a ForeignKey
                                related_obj = related_model.objects.get(product_name=value)  # Assuming 'product_name' is unique
                            else:
                                related_obj = related_model.objects.get(**{field_name: value})
                            data[field_name] = related_obj
                        elif isinstance(field, DecimalField) and value:
                            data[field_name] = float(value)
                    except related_model.DoesNotExist:
                        self.stderr.write(
                            f"Error: {field_name} '{value}' does not exist in {related_model.__name__}. Skipping row.")
                        data[field_name] = None
                    except Exception as e:
                        self.stderr.write(f"Error processing field '{field_name}' in '{model_name}': {e}")
                        data[field_name] = None

                # Determine the unique field for this model
                unique_field = None
                if model_name == "Category":
                    unique_field = 'name'
                elif model_name == "Product":
                    unique_field = 'product_name'
                elif model_name == "Option":
                    unique_field = 'name'

                if not unique_field or unique_field not in data or not data[unique_field]:
                    self.stderr.write(f"Skipping row {row_index} in '{model_name}' due to missing or empty unique identifier.")
                    continue

                # Create or update the model instance
                condition = {unique_field: data[unique_field]}
                try:
                    obj, created = model.objects.update_or_create(defaults=data, **condition)
                    action = 'Created' if created else 'Updated'
                    self.stdout.write(f"{action} {model_name}: {obj}")
                except Exception as e:
                    self.stderr.write(f"Error saving {model_name} (Row {row_index}): {e}")

        self.stdout.write("Data import completed.")
